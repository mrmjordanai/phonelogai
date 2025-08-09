#!/usr/bin/env python3
"""
Excel Processing Service - Handles Excel/Spreadsheet file processing
Supports XLSX and XLS formats with intelligent sheet detection
"""

import json
import sys
import os
import logging
import traceback
import re
from typing import Dict, List, Tuple, Optional, Any, Union
import time

# Excel processing libraries
try:
    import pandas as pd
    import openpyxl
    import xlrd
    import numpy as np
except ImportError as e:
    print(f"Missing Excel dependencies: {e}", file=sys.stderr)
    sys.exit(1)

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Handles Excel file processing with intelligent sheet and data detection"""
    
    def __init__(self):
        self.phone_patterns = [
            r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            r'\b\+?1?[-.]?\d{10}\b',
            r'\(\d{3}\)\s*\d{3}[-.]?\d{4}'
        ]
        
        self.date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}-\d{2}-\d{2}',
            r'\d{1,2}/\d{1,2}/\d{4}'
        ]
        
        self.time_patterns = [
            r'\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AaPp][Mm])?'
        ]
        
        # Common header keywords for phone records
        self.header_keywords = [
            'phone', 'number', 'contact', 'name', 'date', 'time', 
            'duration', 'type', 'direction', 'caller', 'recipient',
            'incoming', 'outgoing', 'missed', 'sms', 'text', 'call'
        ]
    
    def extract_excel_data(self, file_path: str, file_format: str, options: Dict[str, Any]) -> Dict[str, Any]:
        """Extract data from Excel file with intelligent sheet detection"""
        start_time = time.time()
        
        try:
            # Analyze all sheets first
            sheet_info = self._analyze_sheets(file_path, file_format, options)
            
            if not sheet_info:
                raise Exception("No readable sheets found in Excel file")
            
            # Find the best sheet for phone/call data
            best_sheet = self._find_best_sheet(sheet_info)
            
            # Extract data from best sheet
            data = self._extract_sheet_data(file_path, file_format, best_sheet['name'], options)
            
            processing_time = time.time() - start_time
            
            return {
                'success': True,
                'sheets': sheet_info,
                'bestSheet': best_sheet,
                'data': data,
                'totalSheets': len(sheet_info),
                'totalRows': sum(sheet['rowCount'] for sheet in sheet_info),
                'fileFormat': file_format,
                'processingTime': processing_time
            }
            
        except Exception as e:
            logger.error(f"Excel data extraction failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'sheets': [],
                'data': [],
                'totalSheets': 0,
                'totalRows': 0
            }
    
    def _analyze_sheets(self, file_path: str, file_format: str, options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Analyze all sheets in the Excel file"""
        sheets = []
        max_sheets = options.get('maxSheets', 10)
        
        try:
            if file_format == 'xlsx':
                # Use openpyxl for XLSX
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames[:max_sheets]
                
                for i, sheet_name in enumerate(sheet_names):
                    try:
                        worksheet = workbook[sheet_name]
                        sheet_info = self._analyze_xlsx_sheet(worksheet, sheet_name, i, options)
                        if sheet_info:
                            sheets.append(sheet_info)
                    except Exception as e:
                        logger.warning(f"Failed to analyze sheet '{sheet_name}': {e}")
                        continue
                
                workbook.close()
                
            else:  # XLS
                # Use xlrd for XLS (legacy format)
                workbook = xlrd.open_workbook(file_path)
                
                for i, sheet in enumerate(workbook.sheets()[:max_sheets]):
                    try:
                        sheet_info = self._analyze_xls_sheet(sheet, sheet.name, i, options)
                        if sheet_info:
                            sheets.append(sheet_info)
                    except Exception as e:
                        logger.warning(f"Failed to analyze sheet '{sheet.name}': {e}")
                        continue
            
            return sheets
            
        except Exception as e:
            logger.error(f"Sheet analysis failed: {e}")
            return []
    
    def _analyze_xlsx_sheet(self, worksheet, sheet_name: str, index: int, options: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Analyze a single XLSX sheet"""
        try:
            # Get sheet dimensions
            max_row = min(worksheet.max_row or 0, 1000)  # Limit sample size
            max_col = min(worksheet.max_column or 0, 50)
            
            if max_row == 0 or max_col == 0:
                return None
            
            # Extract sample data for analysis
            sample_data = []
            for row in worksheet.iter_rows(min_row=1, max_row=min(max_row, 20), max_col=max_col, values_only=True):
                if any(cell is not None for cell in row):
                    sample_data.append([str(cell) if cell is not None else '' for cell in row])
            
            if not sample_data:
                return None
            
            # Analyze content
            has_headers = self._detect_headers(sample_data)
            data_quality = self._assess_data_quality(sample_data)
            
            return {
                'name': sheet_name,
                'index': index,
                'rowCount': max_row,
                'columnCount': max_col,
                'hasHeaders': has_headers,
                'dataQuality': data_quality,
                'sampleData': sample_data[:5]  # Keep first 5 rows as sample
            }
            
        except Exception as e:
            logger.warning(f"XLSX sheet analysis failed: {e}")
            return None
    
    def _analyze_xls_sheet(self, sheet, sheet_name: str, index: int, options: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Analyze a single XLS sheet"""
        try:
            max_row = min(sheet.nrows, 1000)
            max_col = min(sheet.ncols, 50)
            
            if max_row == 0 or max_col == 0:
                return None
            
            # Extract sample data
            sample_data = []
            for row_idx in range(min(max_row, 20)):
                row_data = []
                for col_idx in range(max_col):
                    try:
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_data.append(str(cell_value) if cell_value else '')
                    except:
                        row_data.append('')
                
                if any(cell for cell in row_data):
                    sample_data.append(row_data)
            
            if not sample_data:
                return None
            
            # Analyze content
            has_headers = self._detect_headers(sample_data)
            data_quality = self._assess_data_quality(sample_data)
            
            return {
                'name': sheet_name,
                'index': index,
                'rowCount': max_row,
                'columnCount': max_col,
                'hasHeaders': has_headers,
                'dataQuality': data_quality,
                'sampleData': sample_data[:5]
            }
            
        except Exception as e:
            logger.warning(f"XLS sheet analysis failed: {e}")
            return None
    
    def _detect_headers(self, sample_data: List[List[str]]) -> bool:
        """Detect if the sheet has headers"""
        if not sample_data or len(sample_data) < 2:
            return False
        
        first_row = sample_data[0]
        
        # Check if first row contains header-like keywords
        header_score = 0
        for cell in first_row:
            if cell and isinstance(cell, str):
                cell_lower = cell.lower()
                for keyword in self.header_keywords:
                    if keyword in cell_lower:
                        header_score += 1
                        break
        
        # Headers likely if >= 30% of columns have header keywords
        return header_score >= max(1, len(first_row) * 0.3)
    
    def _assess_data_quality(self, sample_data: List[List[str]]) -> float:
        """Assess the quality of data in the sheet for phone records"""
        if not sample_data:
            return 0.0
        
        total_cells = sum(len(row) for row in sample_data)
        if total_cells == 0:
            return 0.0
        
        # Count cells with data
        filled_cells = sum(1 for row in sample_data for cell in row if cell and cell.strip())
        fill_ratio = filled_cells / total_cells if total_cells > 0 else 0
        
        # Look for phone/call data patterns
        content_text = ' '.join(' '.join(row) for row in sample_data)
        
        pattern_score = 0
        # Phone numbers
        phone_matches = sum(len(re.findall(pattern, content_text)) for pattern in self.phone_patterns)
        pattern_score += min(0.3, phone_matches / 100.0)
        
        # Dates
        date_matches = sum(len(re.findall(pattern, content_text)) for pattern in self.date_patterns)
        pattern_score += min(0.3, date_matches / 100.0)
        
        # Times
        time_matches = sum(len(re.findall(pattern, content_text)) for pattern in self.time_patterns)
        pattern_score += min(0.2, time_matches / 100.0)
        
        # Call type keywords
        call_keywords = ['incoming', 'outgoing', 'missed', 'call', 'sms', 'text']
        keyword_matches = sum(1 for keyword in call_keywords if keyword.lower() in content_text.lower())
        pattern_score += min(0.2, keyword_matches / 10.0)
        
        return min(1.0, fill_ratio * 0.4 + pattern_score * 0.6)
    
    def _find_best_sheet(self, sheet_info: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Find the sheet most likely to contain phone/call data"""
        if not sheet_info:
            raise Exception("No sheets to analyze")
        
        # Sort by data quality score
        sorted_sheets = sorted(sheet_info, key=lambda x: x['dataQuality'], reverse=True)
        
        # Return the highest quality sheet
        return sorted_sheets[0]
    
    def _extract_sheet_data(self, file_path: str, file_format: str, sheet_name: str, options: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract data from specific sheet"""
        try:
            # Use pandas for data extraction - it handles both formats well
            if file_format == 'xlsx':
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=options.get('headerRow', 0),
                    nrows=options.get('maxRows', 100000)
                )
            else:  # XLS
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=options.get('headerRow', 0),
                    nrows=options.get('maxRows', 100000),
                    engine='xlrd'
                )
            
            # Clean up the dataframe
            if options.get('skipEmptyRows', True):
                df = df.dropna(how='all')
            
            if options.get('trimWhitespace', True):
                # Trim whitespace from string columns
                for col in df.select_dtypes(include=['object']):
                    df[col] = df[col].astype(str).str.strip()
            
            # Convert to list of dictionaries
            records = []
            for index, row in df.iterrows():
                record = {}
                for col in df.columns:
                    value = row[col]
                    
                    # Handle NaN values
                    if pd.isna(value):
                        record[str(col)] = None
                    # Handle numeric types
                    elif isinstance(value, (int, float)):
                        record[str(col)] = value
                    # Handle datetime
                    elif pd.api.types.is_datetime64_any_dtype(type(value)):
                        record[str(col)] = value.isoformat() if hasattr(value, 'isoformat') else str(value)
                    else:
                        record[str(col)] = str(value)
                
                # Only include rows with some data
                if any(v is not None and v != '' for v in record.values()):
                    records.append(record)
            
            return records
            
        except Exception as e:
            logger.error(f"Sheet data extraction failed: {e}")
            return []

def main():
    """Main entry point"""
    if len(sys.argv) != 2:
        print("Usage: python ExcelProcessor.py <json_request>", file=sys.stderr)
        sys.exit(1)
    
    try:
        request = json.loads(sys.argv[1])
        processor = ExcelProcessor()
        
        action = request.get('action')
        file_path = request.get('filePath')
        file_format = request.get('fileFormat', 'xlsx')
        options = request.get('options', {})
        
        if not action or not file_path:
            raise ValueError("Missing required fields: action, filePath")
        
        if not os.path.exists(file_path):
            raise ValueError(f"File not found: {file_path}")
        
        if action == 'extract_excel_data':
            result = processor.extract_excel_data(file_path, file_format, options)
        else:
            raise ValueError(f"Unknown action: {action}")
        
        print(json.dumps(result, default=str))  # default=str handles datetime serialization
        
    except Exception as e:
        logger.error(f"Excel processor error: {e}")
        logger.error(traceback.format_exc())
        
        error_result = {
            'success': False,
            'error': str(e),
            'details': traceback.format_exc()
        }
        
        print(json.dumps(error_result))
        sys.exit(1)

if __name__ == '__main__':
    main()